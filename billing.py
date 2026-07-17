import os
from datetime import datetime
from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.colors import HexColor
from reportlab.pdfgen import canvas
import qrcode


import config

def format_indian_rupee(amount):
    try:
        s = str(int(amount))
    except (ValueError, TypeError):
        return str(amount)
    if len(s) > 3:
        last_3 = s[-3:]
        rest = s[:-3]
        rest = ",".join([rest[max(0, i-2):i] for i in range(len(rest), 0, -2)][::-1])
        return f"{rest},{last_3}"
    return s

def wrap_text(canvas_obj, text, font_name, font_size, max_width):
    """Pixel-based word-wrap. Splits *text* into lines where each line's
    rendered width (via canvas_obj.stringWidth) does not exceed *max_width*
    (in points).  Returns a list of strings."""
    words = text.split()
    if not words:
        return [""]
    lines = []
    current = words[0]
    for word in words[1:]:
        test = current + " " + word
        if canvas_obj.stringWidth(test, font_name, font_size) <= max_width:
            current = test
        else:
            lines.append(current)
            current = word
    lines.append(current)
    return lines


MAROON = HexColor(config.BRAND_MAROON)
GOLD = HexColor(config.BRAND_GOLD)
MAROON_LIGHT = HexColor(config.BRAND_MAROON_LIGHT)

COLS = ["Invoice No", "Customer Name", "Customer Phone", "Item Name",
        "Qty", "Price", "Line Total", "UPI ID", "Status"]


def _col_index(header_row, name):
    for cell in header_row:
        if str(cell.value).strip() == name:
            return cell.column
    raise ValueError(f"Column '{name}' not found in Excel. Check headers.")


def get_pending_invoices():
    """Reads the Excel file and returns pending invoices grouped by Invoice No."""
    wb = load_workbook(config.EXCEL_PATH)
    sheet = wb.active
    header_row = sheet[1]

    idx = {name: _col_index(header_row, name) for name in COLS}

    invoices = {}
    for row in sheet.iter_rows(min_row=2):
        invoice_no = row[idx["Invoice No"] - 1].value
        status = row[idx["Status"] - 1].value
        if not invoice_no:
            continue
        if status and str(status).strip().lower() == "done":
            continue

        invoice_no = str(invoice_no).strip()
        entry = {
            "row_number": row[0].row,
            "customer_name": row[idx["Customer Name"] - 1].value,
            "customer_phone": row[idx["Customer Phone"] - 1].value,
            "item_name": row[idx["Item Name"] - 1].value,
            "qty": row[idx["Qty"] - 1].value or 0,
            "price": row[idx["Price"] - 1].value or 0,
            "upi_id": row[idx["UPI ID"] - 1].value,
        }
        invoices.setdefault(invoice_no, []).append(entry)

    wb.close()
    return invoices


def mark_as_done(invoice_no):
    """Marks all Excel rows belonging to this invoice as Done."""
    wb = load_workbook(config.EXCEL_PATH)
    sheet = wb.active
    header_row = sheet[1]
    idx = {name: _col_index(header_row, name) for name in COLS}

    for row in sheet.iter_rows(min_row=2):
        cell_invoice_no = row[idx["Invoice No"] - 1].value
        if cell_invoice_no and str(cell_invoice_no).strip() == invoice_no:
            row[idx["Status"] - 1].value = "Done"

    wb.save(config.EXCEL_PATH)
    wb.close()


def generate_pdf(invoice_no, items, discount_percent=0, gst_rate=5):
    """Generates a bill-book style branded PDF invoice matching the physical
    Shuhaagan Saree bill book design.  Returns (pdf_path, total_amount)."""
    customer_name = items[0]["customer_name"]
    customer_phone = items[0]["customer_phone"]
    upi_id = items[0]["upi_id"]

    subtotal = sum((it["qty"] or 0) * (it["price"] or 0) for it in items)
    discount_amount = subtotal * (discount_percent / 100)
    taxable = subtotal - discount_amount
    gst_amount = taxable * (gst_rate / 100)
    total_amount = taxable + gst_amount

    qr_path = os.path.join(config.QR_FOLDER, f"{invoice_no}_qr.png")
    if upi_id:
        upi_link = f"upi://pay?pa={upi_id}&pn={config.SHOP_NAME}&am={total_amount}&cu=INR"
        qrcode.make(upi_link).save(qr_path)

    pdf_path = os.path.join(config.PDF_FOLDER, f"{invoice_no}.pdf")
    c = canvas.Canvas(pdf_path, pagesize=A4)
    w, h = A4

    # ── Page margins ──
    margin = 8 * mm
    inner_x = margin + 4 * mm
    inner_w = w - 2 * inner_x
    right_x = w - inner_x

    # ================================================================
    #  WARM GRADIENT BORDER  (yellow → orange → red, layered rects)
    # ================================================================
    border_colors = [
        "#FFD700", "#FFC300", "#FFB000", "#FF9900",
        "#FF7F00", "#FF6600", "#E84B2B", "#C0392B",
    ]
    border_thickness = 8 * mm
    num_bands = len(border_colors)
    band_w = border_thickness / num_bands
    for i, hex_col in enumerate(border_colors):
        offset = i * band_w
        c.setFillColor(HexColor(hex_col))
        c.rect(margin + offset, margin + offset,
               w - 2 * margin - 2 * offset,
               h - 2 * margin - 2 * offset,
               fill=1, stroke=0)
    # White interior
    c.setFillColor(HexColor("#FFFEF5"))
    c.rect(margin + border_thickness, margin + border_thickness,
           w - 2 * margin - 2 * border_thickness,
           h - 2 * margin - 2 * border_thickness,
           fill=1, stroke=0)

    # ================================================================
    #  WATERMARK
    # ================================================================
    c.saveState()
    c.translate(w / 2, h / 2)
    c.rotate(30)
    c.setFont("Times-Bold", 80)
    c.setFillColorRGB(0.95, 0.95, 0.95)  # Faint grey
    c.drawCentredString(0, 0, "SHUHAAGAN SAREE")
    c.restoreState()


    content_top = h - margin - border_thickness - 3 * mm
    content_bot = margin + border_thickness + 3 * mm
    content_left = margin + border_thickness + 3 * mm
    content_right = w - margin - border_thickness - 3 * mm

    # ================================================================
    #  OM GANESHAYA NAMAH  (centered, small, at the very top)
    # ================================================================
    y = content_top - 5 * mm
    c.setFillColor(HexColor("#8B0000"))
    c.setFont("Times-Italic", 10)
    c.drawCentredString(w / 2, y, "|| Shree Ganeshaya Namah ||")

    # ================================================================
    #  HEADER: Logo + Brand Title (left)  |  Contact Block (right)
    # ================================================================
    y -= 10 * mm

    # ── Logo ──
    logo_x = content_left + 4 * mm
    logo_w = 32 * mm
    if os.path.exists(config.LOGO_PATH):
        c.drawImage(config.LOGO_PATH, logo_x, y - 32 * mm,
                     width=logo_w, height=32 * mm,
                     preserveAspectRatio=True, mask="auto")

    # ── Brand title & Contact block with pixel-accurate layout ──
    title_x = logo_x + logo_w + 6 * mm
    contact_right = content_right - 4 * mm

    # ---- Pre-wrap ALL right-column text (pixel-based, once) ----
    addr_font = "Helvetica"
    addr_size = 8
    addr_max_w = 160  # tighter wrap to leave space for title

    addr1_full = f"{config.SHOP_ADDRESSES[0][0]}: {config.SHOP_ADDRESSES[0][1]}"
    addr1_lines = wrap_text(c, addr1_full, addr_font, addr_size, addr_max_w)

    addr2_full = f"{config.SHOP_ADDRESSES[1][0]}: {config.SHOP_ADDRESSES[1][1]}"
    addr2_lines = wrap_text(c, addr2_full, addr_font, addr_size, addr_max_w)

    # Measure every right-column line's pixel width
    right_items = []  # list of (text, width) for debug
    icon_size = 4 * mm
    for p in config.SHOP_PHONES:
        sw = c.stringWidth(p, "Helvetica-Bold", 11) + icon_size + 1.5 * mm
        right_items.append((p + " +icon", sw))
    if hasattr(config, "SHOP_GST") and config.SHOP_GST:
        gst_str = "GST : " + config.SHOP_GST
        sw = c.stringWidth(gst_str, "Helvetica-Bold", 10)
        right_items.append((gst_str, sw))
    for al in addr1_lines:
        sw = c.stringWidth(al, addr_font, addr_size)
        right_items.append((al, sw))
    for al in addr2_lines:
        sw = c.stringWidth(al, addr_font, addr_size)
        right_items.append((al, sw))
    email_sw = c.stringWidth(config.EMAIL, addr_font, addr_size) + icon_size + 1.5 * mm
    right_items.append((config.EMAIL + " +icon", email_sw))
    ig_sw = c.stringWidth(config.INSTAGRAM, addr_font, addr_size) + icon_size + 1.5 * mm
    right_items.append((config.INSTAGRAM + " +icon", ig_sw))

    max_right_w = max(w for _, w in right_items)
    contact_left_bound = contact_right - max_right_w
    allowed_left_w = contact_left_bound - title_x - 5 * mm

    # ---- Pre-wrap categories (pixel-based, once) ----
    cat_size = 7
    cat_full = config.CATEGORIES
    cat_lines = wrap_text(c, cat_full, "Helvetica-Bold", cat_size, allowed_left_w)

    # ---- Auto-scale title fonts until everything fits ----
    title_size = 35
    saree_size = 35
    tagline_size = 12

    while title_size > 14:
        shuhaagan_w = c.stringWidth("SHUHAAGAN ", "Times-Bold", title_size)
        saree_w = c.stringWidth("SAREE", "Times-Bold", saree_size)
        fits = (
            (shuhaagan_w + saree_w) <= allowed_left_w
            and c.stringWidth(config.TAGLINE, "Times-Italic", tagline_size) <= allowed_left_w
        )
        if fits:
            break
        title_size -= 1
        saree_size = title_size
        tagline_size = max(8, tagline_size - 0.5)

    # Re-wrap categories at final allowed_left_w
    cat_lines = wrap_text(c, cat_full, "Helvetica-Bold", cat_size, allowed_left_w)

    # ---- Collect left-column widths for verification ----
    left_items = []
    def _lw(text, font, size):
        sw = c.stringWidth(text, font, size)
        left_items.append((text[:40], title_x, title_x + sw))
        return sw

    shuhaagan_w = c.stringWidth("SHUHAAGAN ", "Times-Bold", title_size)
    saree_w = c.stringWidth("SAREE", "Times-Bold", saree_size)
    left_items.append(("SHUHAAGAN SAREE", title_x, title_x + shuhaagan_w + saree_w))
    _lw(config.TAGLINE, "Times-Italic", tagline_size)
    for cl in cat_lines:
        _lw(cl, "Helvetica-Bold", cat_size)

    max_left_right = max(xe for _, _, xe in left_items)

    # ---- DEBUG: print every line's x-range ----
    print(f"\n=== HEADER DEBUG (allowed_left_w={allowed_left_w:.1f}pt, title={title_size}pt) ===")
    print(f"  Left column  title_x={title_x:.1f}  max_right_edge={max_left_right:.1f}")
    for label, xs, xe in left_items:
        print(f"    L  x=[{xs:.1f} .. {xe:.1f}]  w={xe-xs:.1f}  '{label}'")
    print(f"  Right column  contact_right={contact_right:.1f}  contact_left_bound={contact_left_bound:.1f}")
    for label, rw in right_items:
        rx_start = contact_right - rw
        print(f"    R  x=[{rx_start:.1f} .. {contact_right:.1f}]  w={rw:.1f}  '{label[:50]}'")
    gap = contact_left_bound - max_left_right
    print(f"  GAP between columns: {gap:.1f}pt ({gap/mm:.1f}mm)")
    if gap < 0:
        print("  *** WARNING: columns overlap! ***")
    print("=== END HEADER DEBUG ===\n")

    # ---- Draw left column ----
    c.setFillColor(HexColor("#8B0000"))
    
    # Draw title centered vertically in the available space
    c.setFont("Times-Bold", title_size)
    c.drawString(title_x, y - 6 * mm, "SHUHAAGAN ")
    shuhaagan_w = c.stringWidth("SHUHAAGAN ", "Times-Bold", title_size)
    
    c.setFont("Times-Bold", saree_size)
    c.drawString(title_x + shuhaagan_w, y - 6 * mm, "SAREE")

    tagline_y = y - 6 * mm - (title_size * 0.35) * mm - 1 * mm
    c.setFont("Times-Italic", tagline_size)
    c.drawString(title_x, tagline_y, config.TAGLINE)

    c.setFont("Helvetica-Bold", cat_size)
    c.setFillColor(HexColor("#444444"))
    cat_y = tagline_y - (tagline_size * 0.35) * mm - 2.5 * mm
    for cl in cat_lines:
        c.drawString(title_x, cat_y, cl)
        cat_y -= 4 * mm

    # ---- Draw right column ----
    cy = y + 2 * mm

    c.setFillColor(HexColor("#8B0000"))
    c.setFont("Helvetica-Bold", 11)
    phone_icon_path = os.path.join(config.BASE_DIR, "assets", "phone_icon.png")
    for phone in config.SHOP_PHONES:
        phone_w = c.stringWidth(phone, "Helvetica-Bold", 11)
        if os.path.exists(phone_icon_path):
            c.drawImage(phone_icon_path, contact_right - phone_w - icon_size - 1.5 * mm,
                        cy - 0.5 * mm, width=icon_size, height=icon_size, mask="auto")
        c.drawRightString(contact_right, cy, phone)
        cy -= 5 * mm

    cy -= 1 * mm
    if hasattr(config, "SHOP_GST") and config.SHOP_GST:
        c.setFont("Helvetica-Bold", 10)
        c.drawRightString(contact_right, cy, "GST : " + config.SHOP_GST)
        cy -= 5.5 * mm

    c.setFillColor(HexColor("#444444"))
    c.setFont(addr_font, addr_size)

    # Store 1 — use the pre-wrapped addr1_lines
    for al in addr1_lines:
        c.drawRightString(contact_right, cy, al)
        cy -= 3.5 * mm

    cy -= 2 * mm
    # Store 2 — use the pre-wrapped addr2_lines
    for al in addr2_lines:
        c.drawRightString(contact_right, cy, al)
        cy -= 3.5 * mm

    cy -= 1.5 * mm
    email_w = c.stringWidth(config.EMAIL, addr_font, addr_size)
    email_icon_path = os.path.join(config.BASE_DIR, "assets", "email_icon.png")
    if os.path.exists(email_icon_path):
        c.drawImage(email_icon_path, contact_right - email_w - icon_size - 1.5 * mm,
                    cy - 0.5 * mm, width=icon_size, height=icon_size, mask="auto")
    c.drawRightString(contact_right, cy, config.EMAIL)
    cy -= 4.5 * mm

    ig_text_w = c.stringWidth(config.INSTAGRAM, addr_font, addr_size)
    icon_path = os.path.join(config.BASE_DIR, "assets", "instagram_icon.png")
    if os.path.exists(icon_path):
        c.drawImage(icon_path, contact_right - ig_text_w - icon_size - 1.5 * mm,
                    cy - 0.5 * mm, width=icon_size, height=icon_size, mask="auto")
    c.drawRightString(contact_right, cy, config.INSTAGRAM)

    # ================================================================
    #  BILL FIELDS ROW  (M/s., Bill No., Challan No., Date)
    # ================================================================
    fields_y = y - 48 * mm

    # Light gold background for fields row
    c.setFillColor(HexColor("#FFF8E1"))
    c.rect(content_left, fields_y - 12 * mm,
           content_right - content_left, 16 * mm, fill=1, stroke=0)

    c.setFillColor(HexColor("#222222"))
    c.setFont("Helvetica", 11)

    # M/s. (customer name) — left side
    ms_x = content_left + 4 * mm
    ms_y = fields_y - 2 * mm
    c.drawString(ms_x, ms_y, "M/s.  ")
    
    ms_width = c.stringWidth("M/s.  ", "Helvetica", 11)
    name_phone_str = f"{customer_name or ''}    ({customer_phone or ''})"
    c.drawString(ms_x + ms_width, ms_y, name_phone_str)
    
    # Draw an underline for the customer name and phone
    line_start_x = ms_x + ms_width - 1 * mm
    name_width = c.stringWidth(name_phone_str, "Helvetica", 11)
    # Ensure the line is at least 60mm long even if name is short
    line_end_x = max(line_start_x + name_width + 5 * mm, line_start_x + 60 * mm)
    
    c.setStrokeColor(HexColor("#222222"))
    c.setLineWidth(0.5)
    c.line(line_start_x, ms_y - 1.5 * mm, line_end_x, ms_y - 1.5 * mm)

    # Bill No, Challan No, Date — right side
    bill_fields_x = content_right - 70 * mm
    c.setFont("Helvetica", 10)
    c.drawString(bill_fields_x, fields_y, f"Bill No.: {invoice_no}")
    c.drawString(bill_fields_x, fields_y - 5 * mm, "Challan No.: _______")
    c.drawString(bill_fields_x, fields_y - 10 * mm,
                 f"Date: {datetime.now().strftime('%d-%m-%Y')}")

    # ================================================================
    #  ITEMS TABLE
    # ================================================================
    table_top = fields_y - 22 * mm

    # Column positions — computed from stringWidth to guarantee 2mm padding
    table_font = "Helvetica"
    table_font_size = 11
    col_pad = 2 * mm  # minimum clear space between text edge and grid line

    # Measure worst-case cell values at the table font/size

    hsn_test = "99999999"
    hsn_need = c.stringWidth(hsn_test, table_font, table_font_size) + 2 * col_pad
    hsn_hdr = c.stringWidth("HSN", "Helvetica-Bold", 10) + 2 * col_pad
    hsn_col_w = max(hsn_need, hsn_hdr)

    pcs_test = "999"
    rate_test = "Rs.99,999"
    amt_test = "Rs.99,999"
    pcs_need = c.stringWidth(pcs_test, table_font, table_font_size) + 2 * col_pad
    rate_need = c.stringWidth(rate_test, table_font, table_font_size) + 2 * col_pad
    amt_need = c.stringWidth(amt_test, table_font, table_font_size) + 2 * col_pad

    # Also check header labels at bold 10pt
    pcs_hdr = c.stringWidth("PCS.", "Helvetica-Bold", 10) + 2 * col_pad
    rate_hdr = c.stringWidth("RATE", "Helvetica-Bold", 10) + 2 * col_pad
    amt_hdr = c.stringWidth("AMOUNT", "Helvetica-Bold", 10) + 2 * col_pad

    pcs_col_w = max(pcs_need, pcs_hdr)
    rate_col_w = max(rate_need, rate_hdr)
    amt_col_w = max(amt_need, amt_hdr)

    no_col_w = 16 * mm
    col_no_x = content_left
    col_desc_x = content_left + no_col_w

    # PCS, RATE, AMOUNT are right-anchored from content_right
    col_amt_right = content_right               # AMOUNT right edge = content_right
    col_amt_x = col_amt_right - amt_col_w       # AMOUNT left edge
    col_rate_right = col_amt_x                   # RATE right edge
    col_rate_x = col_rate_right - rate_col_w     # RATE left edge
    col_pcs_right = col_rate_x                   # PCS right edge
    col_pcs_x = col_pcs_right - pcs_col_w        # PCS left edge
    col_hsn_right = col_pcs_x
    col_hsn_x = col_hsn_right - hsn_col_w

    desc_col_w = col_pcs_x - col_desc_x         # DESCRIPTION gets whatever's left

    print(f"Table columns: NO={no_col_w/mm:.1f}mm  DESC={desc_col_w/mm:.1f}mm  "
          f"PCS={pcs_col_w/mm:.1f}mm  RATE={rate_col_w/mm:.1f}mm  AMT={amt_col_w/mm:.1f}mm")

    # Header band — maroon/gold gradient look (solid maroon with gold accents)
    header_h = 12 * mm
    c.setFillColor(MAROON)
    c.rect(content_left, table_top - header_h,
           content_right - content_left, header_h, fill=1, stroke=0)
    # Gold accent line at top of header
    c.setFillColor(GOLD)
    c.rect(content_left, table_top, content_right - content_left, 1.2 * mm,
           fill=1, stroke=0)

    # Header text
    hy = table_top - 8 * mm
    c.setFillColor(HexColor("#FFFFFF"))
    c.setFont("Helvetica-Bold", 10)
    c.drawCentredString(col_no_x + no_col_w / 2, hy, "NO.")
    c.drawString(col_desc_x + 3 * mm, hy, "DESCRIPTION OF GOODS")
    c.drawCentredString(col_hsn_x + hsn_col_w / 2, hy, "HSN")
    c.drawCentredString(col_pcs_x + pcs_col_w / 2, hy, "PCS.")
    c.drawCentredString(col_rate_x + rate_col_w / 2, hy, "RATE")
    c.drawCentredString(col_amt_x + amt_col_w / 2, hy, "AMOUNT")

    # Draw vertical column lines for the table
    row_h = 14 * mm
    # Calculate how many rows fit (reserve space for total + footer)
    footer_reserve = 95 * mm  # space for total row + terms + signatures + QR
    max_table_bottom = content_bot + footer_reserve
    
    total_table_space = table_top - header_h - max_table_bottom
    available_rows = int(total_table_space / row_h)
    
    if available_rows < len(items):
        available_rows = len(items)
        
    actual_table_bottom = table_top - header_h - available_rows * row_h

    # Table body rows
    ry = table_top - header_h
    c.setFont(table_font, table_font_size)
    for i, it in enumerate(items):
        ry -= row_h
        if ry < actual_table_bottom:
            break
        line_total = (it["qty"] or 0) * (it["price"] or 0)

        # Alternating row background
        if i % 2 == 0:
            c.setFillColor(HexColor("#FFFEF5"))
        else:
            c.setFillColor(HexColor("#FFF5E0"))
        c.rect(content_left, ry, content_right - content_left, row_h,
               fill=1, stroke=0)

        c.setFillColor(HexColor("#222222"))
        c.setFont(table_font, table_font_size)
        ty = ry + 4.5 * mm
        c.drawCentredString(col_no_x + no_col_w / 2, ty, str(i + 1))
        c.drawString(col_desc_x + 3 * mm, ty, str(it.get("item_name") or ""))
        c.drawCentredString(col_hsn_x + hsn_col_w / 2, ty, str(it.get("hsn") or ""))
        c.drawCentredString(col_pcs_x + pcs_col_w / 2, ty, str(it.get("qty") or ""))
        
        formatted_rate = format_indian_rupee(it["price"])
        c.drawRightString(col_rate_right - col_pad, ty, f"Rs.{formatted_rate}")
        
        formatted_amt = format_indian_rupee(line_total)
        c.drawRightString(col_amt_right - col_pad, ty, f"Rs.{formatted_amt}")

    # Fill remaining empty rows to match bill-book look
    for j in range(len(items), available_rows):
        ry -= row_h
        if ry < actual_table_bottom:
            break
        if j % 2 == 0:
            c.setFillColor(HexColor("#FFFEF5"))
        else:
            c.setFillColor(HexColor("#FFF5E0"))
        c.rect(content_left, ry, content_right - content_left, row_h,
               fill=1, stroke=0)

    # Draw table grid lines
    c.setStrokeColor(HexColor("#D4A44A"))
    c.setLineWidth(0.4)
    # Horizontal lines
    ty_line = table_top - header_h
    while ty_line >= actual_table_bottom - 0.1 * mm:
        c.line(content_left, ty_line, content_right, ty_line)
        ty_line -= row_h
    # Vertical column separators (full table height)
    full_table_top = table_top + 1.2 * mm
    for vx in [content_left, col_desc_x, col_hsn_x, col_pcs_x, col_rate_x,
               col_amt_x, content_right]:
        c.line(vx, full_table_top, vx, actual_table_bottom)

    # ── "NO RETURN NO EXCHANGE" ──
    nr_y = actual_table_bottom - 8 * mm
    c.setFillColor(MAROON)
    c.setFont("Helvetica-Bold", 12)
    c.drawString(content_left + 4 * mm, nr_y, "NO RETURN NO EXCHANGE")

    # ── TOTAL BREAKDOWN ──
    c.setFillColor(HexColor("#333333"))
    c.setFont("Helvetica", 10)
    
    breakdown_y = nr_y - 8 * mm
    label_x = col_pcs_x
    value_x = col_amt_right - col_pad
    
    if discount_percent > 0 or gst_rate > 0:
        c.drawString(label_x, breakdown_y, "Subtotal:")
        c.drawRightString(value_x, breakdown_y, f"Rs. {format_indian_rupee(subtotal)}")
        breakdown_y -= 5 * mm
        
    if discount_percent > 0:
        c.drawString(label_x, breakdown_y, f"Discount ({discount_percent}%):")
        c.drawRightString(value_x, breakdown_y, f"- Rs. {format_indian_rupee(discount_amount)}")
        breakdown_y -= 5 * mm
        
        c.drawString(label_x, breakdown_y, "Taxable Value:")
        c.drawRightString(value_x, breakdown_y, f"Rs. {format_indian_rupee(taxable)}")
        breakdown_y -= 5 * mm
        
    if gst_rate > 0:
        half_gst = gst_rate / 2
        c.drawString(label_x, breakdown_y, f"CGST ({half_gst}%):")
        c.drawRightString(value_x, breakdown_y, f"Rs. {format_indian_rupee(gst_amount / 2)}")
        breakdown_y -= 5 * mm
        
        c.drawString(label_x, breakdown_y, f"SGST ({half_gst}%):")
        c.drawRightString(value_x, breakdown_y, f"Rs. {format_indian_rupee(gst_amount / 2)}")
        breakdown_y -= 5 * mm
        
    # ── GRAND TOTAL row ──
    total_row_y = breakdown_y - 2 * mm
    c.setFillColor(MAROON)
    c.setFont("Helvetica-Bold", 14)
    formatted_total = format_indian_rupee(total_amount)
    
    c.drawString(label_x, total_row_y, "GRAND TOTAL")
    c.drawRightString(value_x, total_row_y, f"Rs. {formatted_total}")

    # Gold line under total
    c.setStrokeColor(GOLD)
    c.setLineWidth(2)
    c.line(content_left, total_row_y - 4 * mm,
           content_right, total_row_y - 4 * mm)
    
    # Update nr_y so QR code places correctly
    nr_y = total_row_y

    # ================================================================
    #  UPI QR CODE  (left side, beside terms)
    # ================================================================
    qr_section_y = total_row_y - 12 * mm
    qr_bottom = qr_section_y  # track lowest drawn element
    if upi_id and os.path.exists(qr_path):
        qr_size = 40 * mm
        qr_x = content_left + 4 * mm
        qr_y = qr_section_y - qr_size
        c.drawImage(qr_path, qr_x, qr_y,
                     width=qr_size, height=qr_size)
        c.setFillColor(MAROON)
        c.setFont("Helvetica-Bold", 9)
        c.drawCentredString(qr_x + qr_size / 2, qr_y - 5 * mm,
                            "Scan to Pay via UPI")
        c.setFillColor(HexColor("#555555"))
        c.setFont("Helvetica", 8)
        c.drawCentredString(qr_x + qr_size / 2, qr_y - 9 * mm,
                            f"UPI: {upi_id}")
        terms_x = qr_x + qr_size + 12 * mm
        qr_bottom = qr_y - 9 * mm
    else:
        terms_x = content_left + 4 * mm

    # ================================================================
    #  TERMS OF SALE  (fine print, beside or below QR)
    # ================================================================
    terms_y = qr_section_y - 3 * mm
    c.setFillColor(HexColor("#444444"))
    c.setFont("Helvetica-Bold", 9)
    c.drawString(terms_x, terms_y, "Terms of Sale:")
    terms_y -= 5 * mm
    c.setFont("Helvetica", 8)
    for idx_t, term in enumerate(config.TERMS_OF_SALE, 1):
        c.drawString(terms_x, terms_y, f"{idx_t}. {term}")
        terms_y -= 4.5 * mm

    # ================================================================
    #  SIGNATURES  (positioned relative to terms/QR, not page bottom)
    # ================================================================
    sig_y = min(terms_y, qr_bottom) - 4 * mm

    # Receiver's Signature — bottom-left
    c.setFillColor(HexColor("#222222"))
    c.setFont("Helvetica", 10)
    c.drawString(content_left + 4 * mm, sig_y,
                 "Receiver's Signature: _____________________")

    # "For, SHUHAAGAN SAREE" — bottom-right
    c.setFillColor(MAROON)
    c.setFont("Times-Bold", 12)
    c.drawRightString(content_right - 4 * mm, sig_y + 5 * mm, "For,")
    c.setFont("Times-Bold", 18)
    c.drawRightString(content_right - 4 * mm, sig_y - 2 * mm, "SHUHAAGAN")
    c.setFont("Times-Bold", 12)
    c.drawRightString(content_right - 4 * mm, sig_y - 8 * mm, "SAREE")


    # ================================================================
    #  TERMS & CONDITIONS
    # ================================================================
    c.setFillColor(HexColor("#444444"))
    c.setFont("Helvetica-Bold", 9)
    c.drawString(content_left + 4 * mm, content_bot + 24 * mm, "Terms & Conditions:")
    c.setFont("Helvetica", 8)
    c.drawString(content_left + 4 * mm, content_bot + 19 * mm, "1. Goods once sold will not be taken back.")
    c.drawString(content_left + 4 * mm, content_bot + 14 * mm, "2. Exchange allowed within 7 days with bill only.")
    c.drawString(content_left + 4 * mm, content_bot + 9 * mm, "3. Subject to Surat jurisdiction.")

    c.save()
    return pdf_path, total_amount

